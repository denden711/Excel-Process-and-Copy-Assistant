import os
import shutil
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import openpyxl
import logging

# ログの設定
logging.basicConfig(filename='file_processing.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

def get_xlsx_files(directory):
    """指定されたディレクトリ内のxlsxファイルの数字部分のリストを取得する"""
    try:
        files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
        file_info = []
        for f in files:
            try:
                # ファイル名の形式が `x=数字.xlsx` であることを想定
                number = int(f.split('=')[1].split('.')[0])
                file_info.append((number, os.path.join(directory, f)))
            except (ValueError, IndexError) as e:
                logging.error(f"ファイル名の解析中にエラーが発生しました: {f} - {e}")
                continue
        return sorted(file_info, key=lambda x: x[0])
    except Exception as e:
        logging.error(f"ディレクトリの読み取り中にエラーが発生しました: {directory} - {e}")
        messagebox.showerror("エラー", f"ディレクトリの読み取り中にエラーが発生しました: {e}")
        return []

def read_cell_values(file_path, cells):
    """Excelファイルから指定されたセルの値を読み取る"""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        values = []
        for cell in cells:
            if cell:
                value = sheet[cell].value
                values.append(value)
            else:
                values.append(None)
        return values
    except Exception as e:
        logging.error(f"EXCELファイルの読み取り中にエラーが発生しました: {file_path} - {e}")
        return [None] * len(cells)

def write_to_existing_excel(data, excel_path, formula_choice):
    """既存のEXCELファイルにデータを書き込む"""
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        for row_index, row_data in enumerate(data, start=2):
            sheet.cell(row=row_index, column=1).value = row_data[0]  # A列に番号
            sheet.cell(row=row_index, column=2).value = formula_choice.format(i=row_index)  # B列に数式
            for col_index, value in enumerate(row_data[1:], start=3):  # C列から開始
                sheet.cell(row=row_index, column=col_index).value = value

            # 追加する数式
            sheet.cell(row=row_index, column=33).value = f"=O{row_index}/U{row_index}"  # AG列
            sheet.cell(row=row_index, column=34).value = f"=P{row_index}+Q{row_index}+R{row_index}+S{row_index}+T{row_index}+U{row_index}+V{row_index}+X{row_index}+Y{row_index}+Z{row_index}+AA{row_index}+AB{row_index}"  # AH列
            sheet.cell(row=row_index, column=35).value = f"=J{row_index}+K{row_index}+L{row_index}+M{row_index}+N{row_index}"  # AI列
            sheet.cell(row=row_index, column=36).value = f"=U{row_index}+X{row_index}"  # AJ列
            sheet.cell(row=row_index, column=37).value = f"=P{row_index}+Q{row_index}+U{row_index}+V{row_index}+Y{row_index}+Z{row_index}"  # AK列
            sheet.cell(row=row_index, column=38).value = f"=X{row_index}"  # AL列
            sheet.cell(row=row_index, column=39).value = f"=T{row_index}+AA{row_index}"  # AM列
            sheet.cell(row=row_index, column=40).value = f"=R{row_index}+S{row_index}+AB{row_index}"  # AN列
            sheet.cell(row=row_index, column=41).value = f"=J{row_index}+K{row_index}+L{row_index}+M{row_index}+N{row_index}"  # AO列
            sheet.cell(row=row_index, column=42).value = f"=I{row_index}"  # AP列
            sheet.cell(row=row_index, column=43).value = f"=W{row_index}+AC{row_index}"  # AQ列
            sheet.cell(row=row_index, column=44).value = f"=O{row_index}"  # AR列
            sheet.cell(row=row_index, column=45).value = f"=C{row_index}+D{row_index}+E{row_index}+F{row_index}+G{row_index}+H{row_index}"  # AS列

        workbook.save(excel_path)
        messagebox.showinfo("完了", f"Excelファイルが {excel_path} に保存されました。")
        logging.info(f"データがEXCELファイルに正常に書き込まれました: {excel_path}")
    except Exception as e:
        logging.error(f"EXCELファイルへの書き込み中にエラーが発生しました: {excel_path} - {e}")
        messagebox.showerror("エラー", f"EXCELファイルへの書き込み中にエラーが発生しました: {e}")

def select_directory():
    """名前を取得するディレクトリを選択"""
    try:
        directory = filedialog.askdirectory(title="名前を取得するディレクトリを選択")
        if directory:
            directory_label.config(text=f"ディレクトリ: {directory}")
        return directory
    except Exception as e:
        logging.error(f"ディレクトリ選択中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"ディレクトリ選択中にエラーが発生しました: {e}")
        return None

def select_excel_file():
    """元のEXCELファイルを選択"""
    try:
        excel_file = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="元のEXCELファイルを選択")
        if excel_file:
            excel_file_label.config(text=f"元のEXCELファイル: {excel_file}")
        return excel_file
    except Exception as e:
        logging.error(f"EXCELファイル選択中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"EXCELファイル選択中にエラーが発生しました: {e}")
        return None

def get_user_defined_excel_name():
    """新しいEXCELファイルの名前を入力"""
    try:
        user_defined_excel_name = simpledialog.askstring("入力", "コピーする新しいEXCELファイルの名前を入力してください（拡張子なし）:")
        if user_defined_excel_name:
            excel_name_label.config(text=f"新しいEXCELファイル名: {user_defined_excel_name}")
        return user_defined_excel_name
    except Exception as e:
        logging.error(f"新しいEXCELファイル名入力中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"新しいEXCELファイル名入力中にエラーが発生しました: {e}")
        return None

def select_copy_directory():
    """コピー先のディレクトリを選択"""
    try:
        copy_directory = filedialog.askdirectory(title="コピー先のディレクトリを選択")
        if copy_directory:
            copy_directory_label.config(text=f"コピー先のディレクトリ: {copy_directory}")
        return copy_directory
    except Exception as e:
        logging.error(f"コピー先のディレクトリ選択中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"コピー先のディレクトリ選択中にエラーが発生しました: {e}")
        return None

def select_formula():
    """ユーザーが数式を選択（流量別）"""
    try:
        formula_options = [
            "=1/265.240*A{i}*0.1*1000",  # 流量: 2 L/min
            "=1/663.139*A{i}*0.1*1000",  # 流量: 5 L/min
            "=1/1061.038*A{i}*0.1*1000"  # 流量: 8 L/min
        ]
        formula = simpledialog.askstring("数式選択（流量別）", f"使用する数式を選択してください:\n1. {formula_options[0]} (流量: 2 L/min)\n2. {formula_options[1]} (流量: 5 L/min)\n3. {formula_options[2]} (流量: 8 L/min)")
        if formula in ['1', '2', '3']:
            return formula_options[int(formula) - 1]
        else:
            messagebox.showwarning("警告", "無効な選択です。デフォルトの数式を使用します。")
            return formula_options[0]
    except Exception as e:
        logging.error(f"数式選択中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"数式選択中にエラーが発生しました: {e}")
        return "=1/265.240*A{i}*0.1*1000"

def perform_copy_and_write():
    """コピーと書き込みを実行"""
    try:
        directory = select_directory()
        if not directory:
            return
        
        file_info = get_xlsx_files(directory)
        if not file_info:
            return
        
        original_excel_path = select_excel_file()
        if not original_excel_path:
            return
        
        user_defined_excel_name = get_user_defined_excel_name()
        if not user_defined_excel_name:
            return
        
        new_excel_directory = select_copy_directory()
        if not new_excel_directory:
            return
        
        new_excel_path = os.path.join(new_excel_directory, user_defined_excel_name + ".xlsx")
        shutil.copy(original_excel_path, new_excel_path)
        logging.info(f"EXCELファイルがコピーされました: {original_excel_path} -> {new_excel_path}")

        formula_choice = select_formula()

        # 指定されたセルリスト
        cells = ['B49', 'B62', 'B76', 'B91', 'B107', 'B125', 'B156', 'B164', 'B193', 'B219', 'B249', 'B283', 
                 'B482', 'B668', 'B681', 'B709', 'B723', 'B741', 'B758', 'B770', 'B778', 'B800', 'B809', 'B822', 
                 'B842', 'B863', 'B866', None, 'B775']

        data = []
        for number, file_path in file_info:
            row_data = [number] + read_cell_values(file_path, cells)
            data.append(row_data)
        
        write_to_existing_excel(data, new_excel_path, formula_choice)
    except Exception as e:
        logging.error(f"処理中にエラーが発生しました: {e}")
        messagebox.showerror("エラー", f"処理中にエラーが発生しました: {e}")

# GUIの設定
root = tk.Tk()
root.title("XLSXファイル名取得ツール")
root.geometry("600x400")

label = tk.Label(root, text="名前を取得するディレクトリと元のEXCELファイルを選択し、新しいEXCELファイル名を入力して、コピー先のディレクトリを選択します。", wraplength=550)
label.pack(pady=20)

directory_label = tk.Label(root, text="ディレクトリ: 未選択", wraplength=550)
directory_label.pack(pady=5)

excel_file_label = tk.Label(root, text="元のEXCELファイル: 未選択", wraplength=550)
excel_file_label.pack(pady=5)

excel_name_label = tk.Label(root, text="新しいEXCELファイル名: 未入力", wraplength=550)
excel_name_label.pack(pady=5)

copy_directory_label = tk.Label(root, text="コピー先のディレクトリ: 未選択", wraplength=550)
copy_directory_label.pack(pady=5)

button = tk.Button(root, text="ディレクトリとEXCELを選択", command=perform_copy_and_write)
button.pack(pady=20)

root.mainloop()
